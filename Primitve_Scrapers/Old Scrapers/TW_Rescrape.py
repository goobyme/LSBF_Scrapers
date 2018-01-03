import requests
import bs4
import pandas
import threading
import os
import openpyxl

os.chdir('/mnt/c/Users/Liberty SBF/PycharmProjects')
# os.chdir('C:\\Users\\Liberty SBF\\PycharmProjects')
SEARCHPAGE = 'https://www.transwestern.com/bio?email='
THREADCOUNT = 20
employees = []
init_proto_profile_list = []
elist_lock = threading.Lock()
llist_lock = threading.Lock()


def webdl(url):
    """Downloads web-page (using requests rather than urllib) """
    print('Downloading...%s' % url)
    r = requests.get(url)
    try:
        r.raise_for_status()
        return r
    except requests.HTTPError:
        try:
            print('Download failed for %s' % url)
            return None
        except BlockingIOError:
            return None


def excelreader(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name('contacts(1)')
    email_list = []
    for i in range(2, sheet.max_row):
        email = sheet.cell(row=i, column=3)
        email_list.append(email.value)
    return email_list


def personparsing(page):

    e = {}
    headers = {1:'Name', 2:'Title', 3:'Department'}
    i = 1
    soup = bs4.BeautifulSoup(page.text, 'lxml')
    parent_el = soup.find_all('div', {'class': 'twoThirdColumn_bio desktopONLYBio'})
    try:
        elements = parent_el[0].find_all('span')
        for element in elements:
            e[headers[i]] = element.get_text()
            i += 1
    except IndexError:
        return None
    return e


def threadbot(thread_id):

    sublist = []
    while True:
        llist_lock.acquire()
        if len(init_proto_profile_list) > 0:
            try:
                link = init_proto_profile_list[0]
                init_proto_profile_list.remove(link)
            finally:
                llist_lock.release()
            print('Thread %s parsing %s' % (thread_id, link))
            sublist.append(personparsing(webdl(link)))
        else:
            llist_lock.release()
            print('Thread %s completed parsing' % thread_id)
            break

    elist_lock.acquire()
    try:
        global employees
        employees += sublist
        print('Thread %s wrote to list' % thread_id)
    finally:
        elist_lock.release()


def main():
    global init_proto_profile_list
    global employees
    threads = []
    email_list = excelreader('Transwestern.xlsx')
    for email in email_list:
        page = SEARCHPAGE + email
        link_list.append(page)

    for i in range(THREADCOUNT):
        thread = threading.Thread(target=threadbot, args=(i+1, ))

        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()
    to_save = filter(None, employees)
    data_frame = pandas.DataFrame.from_records(to_save)
    data_frame.to_csv('TW_Additional.csv')
    print('Done')


if __name__ == "__main__":
    main()
