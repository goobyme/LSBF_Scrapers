import openpyxl
import re
from pprint import pprint

sf_wb = openpyxl.load_workbook('C:\\Users\\Liberty SBF\\Desktop\\CCIM\\SFaccounts-20180126.xlsx')
sf_sheet = sf_wb.get_sheet_by_name(sf_wb.get_sheet_names()[0])
web_column = tuple(sf_sheet['C:C'])
salesforce_column = tuple(sf_sheet['B:B'])
id_column = tuple(sf_sheet['A:A'])
sf_wb.close()

wb = openpyxl.load_workbook('C:\\Users\\Liberty SBF\\Desktop\\CCIM\\CCIM.xlsx')
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
company_column = tuple(sheet['A:A'])
domain_column = tuple(sheet['C:C'])
wb.close()

"""Check by name"""
name_matches = {}
email_matches = {}
web_regex = re.compile(r"(?<=www.).+(?=/)?")
for prof_numb, company in enumerate(company_column):
    company_text = company.value
    for row_numb, account in enumerate(salesforce_column):
        account_text = account.value
        raw_web_text = web_column[row_numb].value
        web_text = web_regex.findall(raw_web_text)[0]
        if company_text == account_text and company_text not in name_matches:
            name_matches[company_text] = id_column[row_numb].value
        elif domain_column[prof_numb].value == web_text:
            email_matches[web_text] = id_column[row_numb].value


"""Assign ID"""
wb = openpyxl.load_workbook('C:\\Users\\Liberty SBF\\Desktop\\CCIM\\CCIM.xlsx')
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
for row in sheet.iter_rows():
    if row[0].value in name_matches.keys():
        row[1].value = name_matches.get(row[0].value)
        print('Assigned name id {}'.format(row[1].value))
    elif row[2].value in email_matches.keys():
        row[1].value = email_matches.get(row[2].value)
        print('Assigned email id {}'.format(row[1].value))
    else:
        continue

wb.save('CCIM_2.xlsx')
wb.close()


